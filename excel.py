#!/usr/bin/python3

import re
import requests
import json
import os
import xlrd
import xlwt
import openpyxl

path = 'C:\\Users\\ke.liu\\Desktop\\Testlink\\web-tmp.txt'
data = openpyxl.load_workbook('C:\\Users\\ke.liu\\Desktop\\Testlink\\创通RTN510-新功能测试用例.xlsx')
table = data["IP-DCN全量"]
workbook = openpyxl.load_workbook('C:\\Users\\ke.liu\\Desktop\\Testlink\\用例.xlsx')
ws = workbook.create_sheet('IP-DCN')
#nrows 行，ncols 列
# print(table.name,table.nrows,table.ncols)
# print(table.row_values(0),table.col_values(3),table.cell_value(0,0))
# ncols = [5,6]
# for i in ncols:
#     w = open(path, 'w', encoding='UTF-8')
#     for x in table.col_values(i):
#         print(x)
#         w.write(x + '\n')
#     w.close()
#     count = 1
#     line = open(path,encoding='UTF-8')
#     for x in line.readlines():
#         print(x)
#         ws.cell(count,i,x)
#         count += 1
#         workbook.save('C:\\Users\\ke.liu\\Desktop\\Testlink\\用例.xlsx')
#     line.close()

val = table.cell(row=4,column=3).value
print(val)


# rd_sheet = workbook["IP-DCN"]
# rows = rd_sheet.max_row
# for b in table.col_values(2):
#     print(b)
#     for x in range(1,rows):
#         val=rd_sheet.cell(row=x,column=5).value
#         print(val[0:1])
#         if val[0:1] == 1:
#             print("相同")






